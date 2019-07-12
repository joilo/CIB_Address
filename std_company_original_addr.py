import linecache
import re
import xlrd
import openpyxl
from openpyxl.utils import get_column_letter

# 提取路名的信息，并截掉该字段
def extract_road(str_in):
    road = ' '
    road_pattern = re.compile("(.+?[路道])")  # xxx路|道， 取出第一个 路|道 及之前的内容
    road_pattern_1 = re.compile("(.*?[^\d][街巷弄])")  # xxx街|巷|弄,取出第一个 街|巷|弄 及之前的内容，且之前的内容不能有数字
    road_pattern_clear = re.compile(".*?与(.+)") # xx街与xx路，取后面的字段
    road_match = road_pattern.match(str_in)
    road_match_1 = road_pattern_1.match(str_in)
    if road_match:
        road = road_match.group(1)
    elif road_match_1:
        road = road_match_1.group(1)
    else:
        pass
    index = str_in.find(road)
    str_in = str_in[index + len(road):]
    # 去除 路 多余的信息
    road_match_clear = road_pattern_clear.match(road)
    if road_match_clear:
        road = road_match_clear.group(1)
    return str_in,road

# 提取村名的信息，并截掉该字段
def extract_village(str_in):
    village = ' '
    village_pattern = re.compile("(.*?[村乡屯])") # xxx村|乡|屯 取出第一个 村|乡|屯 及之前的内容
    special_pattern = re.compile("(.*?[村乡屯])[路街]") # 特殊情况 新村路 等
    village_pattern_clear = re.compile(".*[号弄区路](.*)") # 去除村前面的多余信息
    village_match = village_pattern.match(str_in)
    special_match = special_pattern.match(str_in)
    if village_match:
        village = village_match.group(1)
    else:
        pass
    if special_match:
        village = ' '
    index = str_in.find(village)
    str_in = str_in[index + len(village):]
    # 去除 村 多余的信息
    village_match_clear = village_pattern_clear.match(village)
    if village_match_clear:
        village = village_match_clear.group(1)
    return str_in,village

# 提取路名和村名
def extract_village_road(str_in):
    roadFirst_pattern = re.compile(".+?[路街道巷].+?[村乡屯]")
    villageFirst_pattern = re.compile(".+?[村乡屯].+?[路街道巷]")
    village = ' '
    road = ' '
    roadFirst_match = roadFirst_pattern.match(str_in)
    villageFirst_match = villageFirst_pattern.match(str_in)
    # 如果路名在村名前面
    if roadFirst_match:
        str_in, road = extract_road(str_in)
        str_in, village = extract_village(str_in)
    # 如果村名在路名前面
    elif villageFirst_match:
        str_in, village = extract_village(str_in)
        str_in, road = extract_road(str_in)
    else:
        # 有 村|乡|屯 关键字,则先进行村的检索
        if '村' in str_in or '乡' in str_in or '屯' in str_in: #or '庄' in str_in or '里' in str_in:
            str_in, village = extract_village(str_in)
        # 进行路的检索
        str_in, road = extract_road(str_in)
    # 去除公交车号被误认为是路的（181路）
    road_pattern_clear = re.compile(".*?\d")
    road_match_clear = road_pattern_clear.match(road)
    if road_match_clear:
        road = ' '
    return village, road

# txt转xlsx
def txt_to_xlsx(filename, outfile):
    fr = open(filename, 'r+', encoding='utf-8-sig')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    row = 0
    for line in fr:
        row += 1
        line = line.strip()
        line = line.split('^')
        col = 0
        for j in range(len(line)):
            col += 1
            ws.cell(column=col, row=row, value=line[j].format(get_column_letter(col)))
    wb.save(outfile)


def std_original_addr(filepath):
    std_file = open(new_txt, 'w+', encoding='utf-8-sig')
    data = xlrd.open_workbook(filepath)
    table = data.sheets()[0]
    nrows = table.nrows
    town_pattern = re.compile("(.+?(镇|街道))[^路]") # 根据关键字 镇|街道 进行检索和分隔，且后面不为路

    parenthesis_pattern = re.compile(".*?\((.*?)\)") # xxx(xx), 取出括号里的内容
    parenthesis_out_pattern = re.compile("(.*?)\(.*?\)") # 截取括号外的内容

    num_left_pattern = re.compile(".+?[路村道街巷镇](.+)") # 取出 路|村|道|街|巷 后面的内容
    num_pattern = re.compile(".*?(\d+)(号|弄|幢|临|栋|组|室|-|$)") # 取出数字，且后面不跟‘米’
    first_ten_pattern = re.compile("(十)[^一二三四五六七八九]")
    first_ten_1_pattern = re.compile("(十)[一二三四五六七八九]")
    middle_ten_pattern = re.compile(".*?[一二三四五六七八九](十)[一二三四五六七八九]")
    last_ten_pattern = re.compile(".*?[一二三四五六七八九](十)[^一二三四五六七八九]")
    count = 0

    for i in range(nrows):
        count = count + 1
        temp = str(table.row_values(i)[4]).strip()
        city = str(table.row_values(i)[5]).strip()
        district = str(table.row_values(i)[6]).strip()
        town_name = ' '   # 镇|街道
        village_name = ' ' # 村|乡|屯
        road_name = ' ' # 路
        num_str = ' ' # 号
        # 如果地址信息不为空
        if not temp.isspace():

            # 去掉区市信息
            if district in temp:
                index = temp.rfind(district)
                temp = temp[index+len(district):]
            elif city in temp:
                index = temp.rfind(city)
                temp = temp[index+len(city):]
            ############################################################

            # temp里的市、区内容已被去除
            # 将剩下的内容分为括号里和括号外
            temp_out_match = parenthesis_out_pattern.match(temp)
            if temp_out_match:
                temp_out = temp_out_match.group(1)
            else:
                temp_out = temp
            parenthesis_match = parenthesis_pattern.match(temp)
            if parenthesis_match:
                temp_in = parenthesis_match.group(1)
            else:
                temp_in = ' '

            # 先处理括号外的字段，取出 镇|街道 的信息，并去除这个字段
            town_match = town_pattern.match(temp_out)
            if town_match:
                town_name = town_match.group(1)
                index = temp_out.find(town_name)
                temp_out = temp_out[index + len(town_name):]

            # 对括号外的字段提取 村 和 路 的信息
            village_name, road_name = extract_village_road(temp_out)

            # 如果村名为空，看能不能括号里提取出村名
            if village_name == ' ' and temp_in != ' ':
                trash,village_name = extract_village(temp_in)
            # 如果路名为空，看能不能怂括号里提取出路名
            if road_name == ' 'and temp_in != ' ':
                trash,road_name = extract_road(temp_in)

            ##############################################################

            # 从temp里取出号
            # 把 路|村|道|街|巷|弄 后面的内容取出来
            num_left = ' '
            num_left_match = num_left_pattern.match(temp)
            if num_left_match:
                num_left = num_left_match.group(1)
            # 取出来的内容进行文字->阿拉伯数字的转换
            if num_left != ' ':
                first_ten = first_ten_pattern.match(num_left)
                first_ten_1 = first_ten_1_pattern.match(num_left)
                middle_ten = middle_ten_pattern.match(num_left)
                last_ten = last_ten_pattern.match(num_left)

                if first_ten:
                    num_left = num_left.replace("十", "10")
                elif first_ten_1:
                    num_left = num_left.replace("十", "1")
                elif middle_ten:
                    num_left = num_left.replace("十", "")
                elif last_ten:
                    num_left = num_left.replace("十", "0")

                num_left = num_left.replace("一", "1")
                num_left = num_left.replace("二", "2")
                num_left = num_left.replace("三", "3")
                num_left = num_left.replace("四", "4")
                num_left = num_left.replace("五", "5")
                num_left = num_left.replace("六", "6")
                num_left = num_left.replace("七", "7")
                num_left = num_left.replace("八", "8")
                num_left = num_left.replace("九", "9")
                num_left = num_left.replace("〇", "0")

                # 转换好以后的字段，取出第一个数字前的所有内容
                match_num = num_pattern.match(num_left)
                if match_num:
                    num_str = match_num.group(1)
        else:
            pass
        print(str(count) + "   " + temp + "              " + town_name + "              " + village_name + "              " + road_name + "           " + num_str)
        std_file.write(str(table.row_values(i)[0]).strip() + '^' + str(table.row_values(i)[1]).strip() + '^' +
                        str(table.row_values(i)[2]).strip() + '^' + str(table.row_values(i)[3]).strip() + '^' +
                        str(table.row_values(i)[4]).strip() + '^' + str(table.row_values(i)[5]).strip() + '^' +
                        str(table.row_values(i)[6]).strip() + '^' + town_name + '^' + village_name + '^' + road_name + '^' + num_str + '\n')
        std_file.flush()
    std_file.close()

def merge(newfile,oldfile):
    merge_file = open(merge_txt, 'w+', encoding='utf-8-sig')
    base = open(newfile, 'r+', encoding='utf-8-sig')
    addition = open(oldfile, 'r+', encoding='utf-8-sig')
    names = []
    for l in addition:
        l = l.strip()
        l = l.split('^')
        comp_name = l[0]
        names.append(comp_name)

    for line in base:
        line = line.strip()
        line = line.split('^')
        name = line[0] #公司名
        num = line[10] #路号
        street = line[9] #路名
        a = num.isspace()
        b = street.isspace()
        # 如果路名或路号为空，则将检查百度接口标准化的地址里有没有此类信息，有的话补充进来
        if a or b:
            if name in names:
                ind = names.index(name)
                cur = linecache.getline(oldfile, ind + 1)
                cur = cur.strip().split('^')
                num_temp = cur[7]
                street_temp = cur[6]
                c = num_temp.isspace()
                d = street_temp.isspace()
                if a and not c:
                    num = num_temp
                if b and not d:
                    street = street_temp

        merge_file.write(line[0].strip() + '^' + line[1].strip() + '^' + line[2].strip() + '^' + line[3].strip() + '^' +
                         line[4].strip() + '^' + line[5].strip() + '^' + line[6].strip() + '^' + line[7].strip() + '^' +
                         line[8].strip() + '^' + street + '^' + num + '\n')
        merge_file.flush()
    merge_file.close()


if __name__ == "__main__":
    global new_txt
    global new_sheet
    global old_sheet
    global file
    global merge_txt
    global merge_sheet
    #city_list = ['安庆']
    # city_list = ['鞍山','包头','宝鸡','保定','北海','北京','本溪','滨州','沧州','昌吉','常德','常州','成都','赤峰','滁州',
    #              '大理','大连','大庆','大同','德阳','东莞','东营','鄂尔多斯','佛山','福州','抚州','赣州','广州','贵阳','桂林',
    #              '哈尔滨','海口','杭州','合肥','菏泽','衡阳','呼和浩特','湖州','淮安','黄石','惠州','吉林','济南','济宁','嘉兴',
    #              '江门','晋城','晋中','九江','昆明','拉萨','兰州','廊坊','乐山','连云港','聊城','临汾','临沂','柳州','龙岩',
    #              '泸州','洛阳','马鞍山','绵阳','南昌','南充','南京','南宁','南平','南通','宁波','宁德','盘锦','平顶山','莆田',
    #              '齐齐哈尔','钦州','青岛','曲靖']
    city_list = ['泉州','日照','三明','三亚','厦门','汕头','上海','绍兴','深圳','沈阳','十堰','石家庄','苏州','台州','太原','泰安',
                 '泰州','唐山','天津','威海','潍坊','渭南','温州','乌鲁木齐','无锡','芜湖','武汉','西安','西宁','咸阳','湘潭','襄阳',
                 '新乡','新余','信阳','邢台','宿迁','徐州','许昌','烟台','延边','盐城','扬州','伊犁','宜宾','宜昌','宜春','义乌','银川',
                 '鹰潭','营口','榆林','岳阳','湛江','漳州','长春','长沙','长治','镇江','郑州','中山','重庆','珠海','株洲','驻马店','遵义','淄博']
    for city_name in city_list:
        new_txt = 'E:\地址\原始地址清洗合并后\%s_公司.txt'%(city_name)
        new_sheet = 'E:\地址\原始地址清洗合并后\%s_公司.xlsx'%(city_name)
        old_txt = 'E:\地址\公司正则后\%s.txt'%(city_name)
        file = 'E:\地址\公司原始地址\%s_公司_地址.xlsx' % (city_name)
        merge_txt = 'E:\地址\原始地址清洗合并后\%s_公司_合并后.txt'%(city_name)
        merge_sheet = 'E:\地址\原始地址清洗合并后\%s_公司_合并后.xlsx'%(city_name)
        std_original_addr(file)
        print("Successfully saved to " + new_txt)
        txt_to_xlsx(new_txt, new_sheet)
        print("Successfully converted to " + new_sheet)
        merge(new_txt,old_txt)
        print("Successfully merged two standard address sheets and saved to " + merge_txt)
        txt_to_xlsx(merge_txt,merge_sheet)
        print("Successfully converted to " + merge_sheet)
        print("-------------------------------------ALL JOBS DONE-----------------------------------")



# if __name__ == "__main__":
#     merge_txt = 'E:\地址\公司正则后\安庆.txt'
#     data = open(merge_txt, 'r+',encoding='utf-8-sig')
#     lines = data.readlines()
#     str = '中兴造船有限公司'
#     name = ['中兴造船有限公司','sdlfhw']
#     print(name.index(str))
#     print(linecache.getline(merge_txt,1))
