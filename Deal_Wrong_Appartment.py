import urllib
import requests
import http.cookiejar
import pymysql
import re
from bs4 import BeautifulSoup
import time
import csv
import random
from urllib.request import urlopen, quote
import json
import xlrd
import traceback
import openpyxl
from openpyxl.utils import get_column_letter
import os

def initial_AK_pond():   #初始化ak 池 0 为有额度 1为额度已经用完 保存成数组格式
    global ak_dic
    ak_dic = {}
    ak_dic = {
        "g3LUhuGRHmgbf2SFwaCF8yAtGW4nezPS": 0,
        "kIzerPbQFAhs01H85XRNvngmXU73RF8L": 0,
        "mKOV9991B7N98C46kUeNRUXScM067Xys": 0,
        "hvZ5O0of19Alfl7HShUWyOVlFiD8WYWG": 0,
        "cq1NkGxVDIyVZSK1xLV2vsfVwOu5ajhV": 0,
        "5tlPS97zCcgTjbk6gy6AnLVG4p2jtg7u": 0,
        "UKg9gDjYcHMB5hSFMi1HxuQz18f041t1": 0,
        "GEC7Zek74HysO1AKCx1iG6bOXCzTWE6z": 0,
        "9HnBVwKEC01DMgxmINhOSGMt5q1M8kyr": 0,
        "SB3KV3mGWLQ3ncHEk7QfiRNCYHFMYtav": 0,
        "CDjmH9V1ZFfhv9qX2KzGvCrf8UVdUu99": 0,
        "y0j1hsIBRInpjyXub9dFLwsHjnTx73m4": 0,
        "YrwNoo8bNA2Nzfj7pldFXaXVz7iyEPXZ": 0,
        "jGDQ054Yx7n9MNFewodPgykG9UlvlYNa": 0,
        "c5AA39AFSpEAtDCRWWCRhoG5htUrUWvD": 0,
        "lN8FZ9Y8dXeGVdv7aTKvdcnkpXUuMcpQ": 0,
        "zer4hmUsf2Cppl2Z3ozkRMrGx6phGMVf": 0,
        "KdaCBLpAZrUApkiVqjYFSheusOwf2bhh": 0,
        "gwThbIBPPOlUYBQIMUhIP5haNLLkG3Nx": 0,
        "yHHqlqI0pLycBZVlMRjCFtQ8HSyMVWgv": 0,
        "cON4dGCQZyz5IqglB7dYcbCPyrYxoxu1": 0,
    }

class NoneAKException(Exception):
    def __init__(self,message):
        Exception.__init__(self)
        self.message=message

def mkdir(path):    #用于创建相对路径下的文件夹
    # os.makedirs(path)
    isExists=os.path.exists(path)

    if not isExists:
        # 如果不存在则创建目录
    # 创建目录操作函数
        os.makedirs(path)

        print(path + ' 创建成功')
        return True
    else:
        # 如果目录存在则不创建，并提示目录已存在
        print(path + ' 目录已存在')
        return False

def exchange_AK():
    for line in ak_dic.items():
        if line[1] == 0:
            return line[0]
    print("ak池的额度全部用完了")
    return None
def list_name(keyname, value1, dict1=None):  # 将一条数据压缩成字典格式
    dict1 = dict(zip(keyname, value1))
    return dict1

def getlnglat(address):   #调用百度地图API对于酒店地址信息进行经纬度逆向解析,传入address字符串
    url = 'http://api.map.baidu.com/geocoder/v2/'
    output = 'json'
    ak = 'yHHqlqI0pLycBZVlMRjCFtQ8HSyMVWgv'
    add = quote(address) #由于本文城市变量为中文，为防止乱码，先用quote进行编码
    url_send = url + '?' + 'address=' + add  + '&output=' + output + '&ak=' + ak
    req = urlopen(url_send)
    res = req.read().decode() #将其他编码的字符串解码成unicode
    temp = json.loads(res) #对json数据进行解析
    lng = temp['result']['location']['lng']
    lat = temp['result']['location']['lat']
    return lng, lat

def reverseLng(lng, lat,ak):   #经纬度反向解析, 经度在前 纬度在后  获得标准地址库所用json文件
    add_list = []
    url = "http://api.map.baidu.com/geocoder/v2/"
    output = 'json'
    ak = ak
    url_send = url + "?callback=renderReverse&location=%s,%s&output=json&pois=1&latest_admin=1&ak=%s&radius=300" % (lat, lng, ak)
    req = urlopen(url_send)
    res = req.read().decode()  # 将其他编码的字符串解码成unicode
    if str(res[:29]) == 'renderReverse&&renderReverse(':
        temp = json.loads(res[29:-1])
        city = temp['result']['addressComponent']['city']
        district = temp['result']['addressComponent']['district']
        formatted_address = temp['result']['formatted_address']
        pois = temp['result']['pois']
        street = temp['result']['addressComponent']['street']  # 获取小区所在街道 用于建立标准路库
        for line in pois:
            model = []
            model.append(line['addr'])
            model.append(line['distance'])
            model.append(line['name'])
            lng = str(line['point']['x'])
            lat = str(line['point']['y'])
            model.append(lat)
            model.append(lng)
            add_list.append(model)
        return add_list, city, district, formatted_address, street
    else:
        temp = json.loads(res)
        if temp['status'] == 301 or temp['status'] == 302 or temp['status'] == 401 or temp['status'] == 402:
            ak_dic[ak] = 1  # 将当前AK的状态设置为已经跑完  P.S 1为已经跑完 0 为还有剩余额度
            print("捕获到AK额度不够的异常")
            for i in range(len(ak_dic)):
                ak_dic[ak] = 1
                ak = exchange_AK()  # 换一个AK
                print("已经更换AK", ak)
                if ak == None:  # 如果调用ak 之后为None 证明ak池的额度全部用完 错误文件记录当前运行结束时的状态
                    print("配额全部用完啦！")
                    raise NoneAKException("AK用完了")
                print("-----------------等待3s-------------------")
                time.sleep(3)
                reverseLng(lng, lat, ak)

def reverseLng1(name,lng, lat,apartment,ak, city):   #经纬度反向解析   经度在前 纬度在后
    add_list = []
    url = "http://api.map.baidu.com/geocoder/v2/"
    output = 'json'
    ak1 = ak
    url_send = url + "?callback=renderReverse&location=%s,%s&output=json&pois=1&latest_admin=1&ak=%s" % (lat, lng, ak1)
    req = urlopen(url_send)
    res = req.read().decode()  # 将其他编码的字符串解码成unicode
    if str(res[:29]) == 'renderReverse&&renderReverse(':
        temp = json.loads(res[29:-1])
        district = temp['result']['addressComponent'].get('district',' ')
        formatted_address = temp['result'].get('formatted_address',' ')
        street = temp['result']['addressComponent'].get('street',' ')  #获取小区所在街道 用于建立标准路库
        #city = temp['result']['addressComponent'].get('city','')
        road_file.write(apartment + "^" + name + "^" + city + "^" + district + "^" + street + "^" + formatted_address + "^" + lat + "^" + lng + "\n")  # 写入文件
        print(street+"已完成")
        road_file.flush()
    else:
        temp = json.loads(res)
        if temp['status'] == 301 or temp['status'] == 302 or temp['status'] == 401 or temp['status'] == 402:
            ak_dic[ak1] = 1  # 将当前AK的状态设置为已经跑完  P.S 1为已经跑完 0 为还有剩余额度
            print("捕获到AK额度不够的异常")
            for i in range(len(ak_dic)):
                ak1 = exchange_AK()  # 换一个AK
                ak_dic[ak1] = 1
                print("已经更换AK", ak1)
                if ak1 == None:  # 如果调用ak 之后为None 证明ak池的额度全部用完 错误文件记录当前运行结束时的状态
                    print("配额全部用完啦！")
                    raise NoneAKException("AK用完了")
                print("-----------------等待3s-------------------")
                time.sleep(3)
                url_send = url + "?callback=renderReverse&location=%s,%s&output=json&pois=1&latest_admin=1&ak=%s" % (lat, lng, ak1)
                req = urlopen(url_send,timeout=1)
                res = req.read().decode()  # 将其他编码的字符串解码成unicode
                if str(res[:29]) == 'renderReverse&&renderReverse(':
                    temp = json.loads(res[29:-1])
                    district = temp['result']['addressComponent'].get('district',' ')
                    formatted_address = temp['result'].get('formatted_address',' ')
                    street = temp['result']['addressComponent'].get('street',' ')  #获取小区所在街道 用于建立标准路库
                    #city = temp['result']['addressComponent'].get('city','')
                    road_file.write(apartment+"^"+name+"^"+city+"^"+ district + "^" + street+"^"+formatted_address+"^"+lat+"^"+lng+"\n")  #写入文件
                    print(street+"已完成")
                    road_file.flush()
                    break
    return ak1


def BaiduAPI_singleSearch(key, region, ak):
    key_encode = quote(key)
    district = quote(region)
    url = "http://api.map.baidu.com/place/v2/search"
    url_send = url + "?query={}&region={}&output=json&ak={}".format(key_encode, district, ak)
    req = urlopen(url_send, timeout=60)
    res = req.read().decode()
    temp = json.loads(res)
    if temp['status'] == 0:
        name = key
        result = temp['results']
        for line in result:
            subname = line.get('name', '')
            city = line.get('city', '')
            area = line.get('area', '')
            addr = line.get('address', '')
            if 'location' in line:
                lat = line['location'].get('lat', '')
                lng = line['location'].get('lng', '')
            else:
                lat = ''
                lng = ''
            data = str(name) + "^" + str(subname) + "^" + str(city) + "^" + str(area) + "^" + str(addr) + "^" + str(
                lat) + "^" + str(lng)
            new_txt.write(data + "\n")  # 写入txt
            new_txt.flush()
    elif temp['status'] == 301 or temp['status'] == 302 or temp['status'] == 401 or temp['status'] == 402:
        ak_dic[ak] = 1  # 将当前AK的状态设置为已经跑完  P.S 1为已经跑完 0 为还有剩余额度
        print("捕获到AK额度不够的异常")
        ak = exchange_AK()  # 换一个AK
        print("已经更换AK", ak)
        if ak == None:  # 如果调用ak 之后为None 证明ak池的额度全部用完 错误文件记录当前运行结束时的状态
            print("配额全部用完啦！")
            raise NoneAKException("AK用完了")
        print("-----------------等待3s-------------------")
        time.sleep(3)
        BaiduAPI_singleSearch(key, region, ak)
    else:
        pass
    return ak

#将xlsx文件按照一个一个sheet 区分开
def read_error_file(road):
    '''
    :param road:
    :return:
    sheet_names: 所有sheet的名字（即城市名）组成的列表
    data:xlrd访问的对象
    '''
    data = xlrd.open_workbook(road)
    sum =0
    sheet_names = []
    for sheet in data.sheets():
        sheet_names.append(sheet.name)
    return sheet_names,data

#将xlsx文件的一个sheet 读取成错误小区组成的列表 例 上海XX小区
def read_error_sheet(sheet_name,data):
    appartment_list = []
    table = data.sheet_by_name(sheet_name)
    nrows = table.nrows
    for i in range(nrows):
        string = sheet_name + table.row_values(i)[0]
        appartment_list.append(string)
    return list(set(appartment_list)) # 用集合来去重

#将一个sheet当中 也就是一个城市所有出问题的进行处理 最终生成该城市的路库 json文件
def deal_one_sheet(city,appartment_list,ak):
    error_road = road_base + "%s/%s_error.txt" % (city,city)
    road = road_base + "%s/%s.txt" % (city,city)
    error_file = open(error_road, 'a+')
    fileObject = open(road, 'w+')
    start_string = '{"result":['
    fileObject.write(start_string)
    sum = 0  #计数器 用于判断是否要加 ，（逗号）
    lengenth = len(appartment_list)  #数组长度
    for line in appartment_list:
        arr = {}
        arr['name'] = str(line).replace(city,'')
        print(arr)
        lng1, lat1 = getlnglat(line)
        arr['lng'] = lng1
        arr['lat'] = lat1
        try:
            add_list, city, district, formatted_address, street = reverseLng(lng1, lat1, ak)
        except Exception as e:
            print("百度地图API没解析出来 ", line, "已写入错误文件")
            error_file.write(line)
            error_file.write("\n")
            error_file.flush()
            continue
        arr['formatted_address'] = formatted_address
        arr['city'] = city
        arr['district'] = district
        arr['street'] = street
        key2 = ['add', 'distance', 'name', 'lat', 'lng']
        add_save_list = []
        for add in add_list:
            a2 = {}
            arr2 = list_name(key2, add, a2)
            arr2['count'] = sum
            add_save_list.append(arr2)
        arr['add_list'] = add_save_list  # 将该小区300m范围内路库信息加入json串中
        arr = json.dumps(arr, ensure_ascii=False)
        fileObject.write(arr)
        print(line, "处理成功")
        if sum < lengenth - 1:  # 如果不是最后一行数据，两条json之间要加 ","  如果是最后一行 不要加 ","
            fileObject.write(",")
            fileObject.write('\n')
            sum += 1
        else:
            sum += 1
    fileObject.flush()
    fileObject.write(']}')
    fileObject.close()

def read_Appartment_Info(ak, road):#读取某个路库文件（json）中所有小区信息 并再次操作 进行标准化
    '''
    :param ak:  百度API AK
    :param road:  存储路库的地址  例如 D://working//Shanghai_error.txt
    :return:
    '''
    file = open(road, encoding='gbk')
    list_lat_lnt=[]
    global save_plot
    road_split = road.split("/")
    road2 = ""
    for i in road_split[:-1]:
        road2 += i + "/"
    city = str(road_split[-1:][0][:-4])   # 获取在文件路径中的城市名
    print(city)
    save_road = road_base + "%s/%s_已爬取.txt" % (city,city)
    save_plot = open(save_road, 'a+', encoding='utf-8-sig')
    save_plot.seek(0,0)
    global road_file
    road_file_road = road_base + "%s/%s_小区_经纬度错误.txt" % (city,city)
    road_file = open(road_file_road, 'a+', encoding='utf-8-sig')
    global error_list
    road_file_error_road = road_base + "%s/%s_小区_经纬度错误_error.txt" % (city,city)
    error_list = open(road_file_error_road, 'a+', encoding='utf-8-sig')
    list_lat_lnt=[line.replace("\n", "") for line in save_plot]
    print(list_lat_lnt)
    data = json.load(file)
    for line in data['result']:
        lat_lnt = str(line['lng']).strip()+","+str(line['lat']).strip()+","+line['name'].strip()
        print(lat_lnt)
        if lat_lnt in list_lat_lnt:
            print(lat_lnt, "已经搜索过")
            continue
        try:
            #reverseLng(name,line['lng'],line['lat'],ak)
            road_file.write(line['name'].strip()+"^"+line['name'].strip()+"^"+city+"^"+line['district'].strip()+ "^" +line['street'].strip()+"^"+line['formatted_address'].strip()+"^"+str(line['lat']).strip()+"^"+str(line['lng']).strip()+"\n")
            road_file.flush()
            for i in line['add_list']:
                ak=reverseLng1(i['name'].strip(),i['lng'].strip(), i['lat'].strip(),line['name'].strip(), ak = ak, city=city)
            save_plot.write(lat_lnt+"\n")
            save_plot.flush()
        except NoneAKException as e:  # 捕捉AK额度不够的异常
            print("AK额度不够啦")
            break
        except Exception as e:
            print(e)
            error_list.write("其他异常:爬取区域为"+lat_lnt+"\n")
            error_list.write(traceback.format_exc()+"\n")
            error_list.flush()
    error_list.close()
    road_file.close()
    save_plot.close()

def deal_all_road_base(sheet_names,ak):  #访问所有小区路库  将其存储标准化
    '''
    :param sheet_names: 传入存储着所有城市名字的城市列表
    :param ak: 百度地图AK
    :return:
    '''
    for city in sheet_names:
        print(city)
        app_road = road_base + city + "/" + city + ".txt"
        read_Appartment_Info(ak = ak,road = app_road)
        print(city, "小区路库标准化完成")
    print("==========================所有城市小区标准化已完成==========================")
    print("==========================等待进行Excel的转换==============================")
    time.sleep(5)

def txt_to_xlsx(filename, outfile):  #将txt 文件 转化成 Excel文件
    '''
    :param filename:  需要转化的txt文件的绝对路径
    :param outfile:   生成的Excel文件的绝对路径
    :return:
    '''
    fr = open(filename, 'r+', encoding='utf8')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    row = 0
    for line in fr:
        row += 1
        line = line.strip()
        line = line.split('^')
        col = 0
        for j in range(len(line)):
            col += 1
            ws.cell(column=col, row=row, value=line[j].format(get_column_letter(col)))
    wb.save(outfile)

def cut_standard_add(xlsx_file, output_file):   #用于处理一个xlsx文件中的小区信息 并对其进行切分
    '''
    :param xlsx_file:  输入的Excel文件小区信息（未切分）
    :param output_file:  输出的txt文件（已切分）
    :return:
    '''
    fileObject = open(output_file, 'w', encoding='utf8')
    data = xlrd.open_workbook(xlsx_file)
    table = data.sheets()[0]
    nrows = table.nrows
    road_pattern = re.compile(".*?[^路]\((.*?)\)")
    road_pattern_1 = re.compile("(.*?(路|大道|街))")
    road_pattern_2 = re.compile("(.+)")
    num_leftpattern = re.compile(".*?(路|大道|街)(.+)")
    num_pattern = re.compile(".*?(\d+)")
    first_ten_pattern = re.compile("(十)[^一二三四五六七八九]")
    first_ten_1_pattern = re.compile("(十)[一二三四五六七八九]")
    middle_ten_pattern = re.compile(".*?[一二三四五六七八九](十)[一二三四五六七八九]")
    last_ten_pattern = re.compile(".*?[一二三四五六七八九](十)[^一二三四五六七八九]")
    for i in range(nrows):
        temp = str(table.row_values(i)[4]).strip()
        road_name = ' '
        road_match = road_pattern.match(temp)
        if road_match:
            road_name = road_match.group(1)
        else:
            road_match_1 = road_pattern_1.match(temp)
            if road_match_1:
                road_name = road_match_1.group(1)
            else:
                road_match_2 = road_pattern_2.match(temp)
                if road_match_2:
                    road_name = road_match_2.group(1)
        temp = str(table.row_values(i)[5]).strip()
        num_left = ' '
        num_leftmatch = num_leftpattern.match(temp)
        if num_leftmatch:
            num_left = num_leftmatch.group(2)
        num_str = ' '
        if num_left != ' ':
            first_ten = first_ten_pattern.match(num_left)
            first_ten_1 = first_ten_1_pattern.match(num_left)
            middle_ten = middle_ten_pattern.match(num_left)
            last_ten = last_ten_pattern.match(num_left)

            if first_ten:
                num_left = num_left.replace("十", "10")
            elif first_ten_1:
                num_left = num_left.replace("十", "1")
            elif middle_ten:
                num_left = num_left.replace("十", "")
            elif last_ten:
                num_left = num_left.replace("十", "0")
            num_left = num_left.replace("一", "1")
            num_left = num_left.replace("二", "2")
            num_left = num_left.replace("三", "3")
            num_left = num_left.replace("四", "4")
            num_left = num_left.replace("五", "5")
            num_left = num_left.replace("六", "6")
            num_left = num_left.replace("七", "7")
            num_left = num_left.replace("八", "8")
            num_left = num_left.replace("九", "9")
            match_num = num_pattern.match(num_left)
            if match_num:
                num_str = match_num.group(1)
        fileObject.write(str(table.row_values(i)[0]).strip() + '^' + str(table.row_values(i)[1]).strip() + '^' +
                        str(table.row_values(i)[2]).strip() + '^' + str(table.row_values(i)[3]).strip() + '^' +
                        str(table.row_values(i)[5]).strip() + '^' + str(table.row_values(i)[6]).strip() + '^' +
                        str(table.row_values(i)[7]).strip() + '^' + road_name + '^' + num_str + '\n')
    fileObject.close()

#用于切分所有城市小区的地址
def cut_to_standard(sheet_names):
    '''
    :param sheet_names: 传入所有小区的城市列表
    :return:
    '''
    for sheet in sheet_names:
        outfile_road = road_base + "%s/%s_小区_经纬度错误_切分.txt" % (sheet,sheet)
        xlsx_file = road_base + "%s/%s_小区_经纬度错误.xlsx" % (sheet,sheet)
        cut_standard_add(xlsx_file,outfile_road)
        print(sheet, "切分已完成")
    print("==========================所有切分已完成=============================")
    print("==========================等待第三次迭代=============================")
    time.sleep(5)

def deal_lack_data(filepath, city):
    global new_txt
    new_txt_road = road_base + "%s/%s_小区_缺失.txt" % (city,city)
    new_txt = open(new_txt_road, 'a+' , encoding = 'utf-8')
    searched_list_road = road_base + "%s/%s_小区_searched.txt" % (city,city)
    searched_list = open(searched_list_road, 'a+', encoding='utf-8-sig')
    searched_list.seek(0,0)
    searched_list1=[line.encode('utf8').decode('utf-8-sig').replace("\n","") for line in searched_list]
    error_list_road = road_base + "%s/%s_小区_报错.txt" % (city,city)
    error_list3 = open(error_list_road, 'a+', encoding='utf-8')
    ak = exchange_AK()
    data = xlrd.open_workbook(filepath)
    table = data.sheets()[0]
    nrows = table.nrows
    nonstandard_addr_count = 0
    success_count = 0
    print("Finish reading xlsx")
    for i in range(nrows):
        num = table.row_values(i)[8]  # 路号
        street = table.row_values(i)[7]  # 路名
        print(num, street)
        if str(num).isspace() or str(street).isspace():  # 如果路名或路号为空，则将名字和行政区域输入百度地图，返回相关条目作为此地址的补充
            nonstandard_addr_count = nonstandard_addr_count + 1
            keyword = table.row_values(i)[0]  # 公司/小区名字
            lat = table.row_values(i)[4]
            lng = table.row_values(i)[5]
            addr = table.row_values(i)[3]
            cur = str(keyword) + "^" + str(addr) + "^" + str(lat) + "^" + str(lng)
            # 如果是已爬取的信息，跳过
            if cur in searched_list1:
                continue
            district = table.row_values(i)[2]  # 区名
            try:
                ak = BaiduAPI_singleSearch(keyword, district, ak)
                success_count = success_count + 1
                print("Finish  " + str(success_count) + " " + str(keyword))
                searched_list.write(cur + "\n")  # 写入txt
                searched_list.flush()
            except Exception as e:
                print(e)
                error_list3.write(cur + "\n")  # 写入txt
                error_list3.flush()
                continue
        else:
            print("未缺失路名或门牌号，跳过...")
    print("============================================================================")
    print("Number of nonstandard address: " + str(nonstandard_addr_count))
    print("Number of successful search in BaiduAPI: " + str(success_count))
    print("============================================================================")

def run(road,root_road, ak):   #用于从头至尾将小区信息完善
    sheet_names, data = read_error_file(road)
    global road_base  #所有文件存储的根目录
    road_base = root_road

    for sheet in sheet_names: #将出错小区 名字 和 城市结合 重新放入百度API 进行经纬度的抓取 再逆向解析 生成最终的所有城市路库文件
        path = './小区经纬度错误2/%s'% (sheet) #创建该城市的文件夹
        mkdir(path)
        appartment_list = read_error_sheet(sheet,data)
        deal_one_sheet(sheet,appartment_list,ak)
    print("=====================json格式路库建设已完成===========================")
    time.sleep(5)

    deal_all_road_base(sheet_names,ak)  #第二次迭代 将300mPOI中地址标准化用于扩充路库
    print("=====================标准路库建设已经完成=============================")

    for sheet in sheet_names:  #用于将所有路库的txt格式转化成xlsx格式
        road = road_base + "%s/%s_小区_经纬度错误.txt" % (sheet,sheet)
        outroad = road_base + "%s/%s_小区_经纬度错误.xlsx" % (sheet,sheet)
        txt_to_xlsx(road,outroad)
        print(sheet, "已完成")
    print("=====================所有标准路库转化Excel已完成============================")

    cut_to_standard(sheet_names)  #将标准路库中的地址信息进行切分 保存到txt文件

    for sheet in sheet_names:  #用于将切分好的路库的txt格式转化成xlsx格式
        road = road_base + "%s/%s_小区_经纬度错误_切分.txt" % (sheet,sheet)
        outroad = road_base + "%s/%s_小区_经纬度错误_切分.xlsx" % (sheet,sheet)
        txt_to_xlsx(road,outroad)
        print(sheet, "已完成")
    print("=====================所有标准路库切分已经完成==========================")
    time.sleep(5)
    #
    for sheet in sheet_names:   #检查切分好的路库 对于路名或门牌号缺少的进行第三次迭代
        city = sheet
        xlsx_file = root_road + "%s/%s_小区_经纬度错误_切分.xlsx" % (city, city)  #这里文件名换成整合切分好了的小区文件
        deal_lack_data(xlsx_file,city)
    print("=====================第三次迭代结束====================================")


    for sheet in sheet_names:
        road = road_base + "%s/%s_小区_缺失.txt" % (sheet,sheet)
        outroad = road_base + "%s/%s_小区_标准库.xlsx" % (sheet,sheet)
        txt_to_xlsx(road,outroad)
        print(sheet, "已完成")
    print("=====================所有标准路库切分已经完成==========================")

if __name__ == "__main__":
    initial_AK_pond()
    ak = exchange_AK()
    road = r"C:\Users\刘凡\Desktop\小区经纬度错误(1).xlsx"  #存储出错小区的xlsx 一个sheet为一个城市
    root_road = "./小区经纬度错误2/"  #所有文件存储的根目录
    run(road, root_road, ak)

