# -*- coding:utf-8 -*-
import json
import sys
import openpyxl
import xlrd
from openpyxl.utils import get_column_letter
#import urllib
from urllib.request import urlopen, quote
# import xlrdFileExistsError: [Errno 17]
import time
import os

class NoneAKException(Exception):
    def __init__(self,message):
        Exception.__init__(self)
        self.message=message

def initial_AK_pond():  # 初始化ak 池 0 为有额度 1为额度已经用完 保存成数组格式
    global ak_dic
    ak_dic = {}
    ak_dic = {
        "5tlPS97zCcgTjbk6gy6AnLVG4p2jtg7u": 0,
        "UKg9gDjYcHMB5hSFMi1HxuQz18f041t1": 0,
        "GEC7Zek74HysO1AKCx1iG6bOXCzTWE6z": 0,
        "9HnBVwKEC01DMgxmINhOSGMt5q1M8kyr": 0,
        "SB3KV3mGWLQ3ncHEk7QfiRNCYHFMYtav": 0,
    }

def exchange_AK():
    for line in ak_dic.items():
        if line[1] == 0:
            return line[0]
        else:
            continue
    print("ak池的额度全部用完了")
    return None

def BaiduAPI_singleSearch(key, region, ak):
    key_encode = quote(key)
    district = quote(region)
    url = "http://api.map.baidu.com/place/v2/search"
    url_send = url + "?query={}&region={}&output=json&ak={}".format(key_encode, district, ak)
    req = urlopen(url_send, timeout=60)
    res = req.read().decode()
    temp = json.loads(res)
    if temp['status'] == 0:
        name = key
        result = temp['results']
        for line in result:
            subname = line.get('name', '')
            city = line.get('city', '')
            area = line.get('area', '')
            addr = line.get('address', '')
            if 'location' in line:
                lat = line['location'].get('lat', '')
                lng = line['location'].get('lng', '')
            else:
                lat = ''
                lng = ''
            data = str(name) + "^" + str(subname) + "^" + str(city) + "^" + str(area) + "^" + str(addr) + "^" + str(
                lat) + "^" + str(lng)
            new_txt.write(data + "\n")  # 写入txt
            #new_txt.flush()
    elif temp['status'] == 301 or temp['status'] == 302 or temp['status'] == 401 or temp['status'] == 402:
        ak_dic[ak] = 1  # 将当前AK的状态设置为已经跑完  P.S 1为已经跑完 0 为还有剩余额度
        print("捕获到AK额度不够的异常")
        ak = exchange_AK()  # 换一个AK
        print("已经更换AK", ak)
        if ak == None:  # 如果调用ak 之后为None 证明ak池的额度全部用完 错误文件记录当前运行结束时的状态
            print("配额全部用完啦！")
            raise NoneAKException("AK用完了")
        print("-----------------等待3s-------------------")
        time.sleep(3)
        BaiduAPI_singleSearch(key, region, ak)
    else:
        pass
    return ak


def run(filepath):
    ak = exchange_AK()
    data = xlrd.open_workbook(filepath)
    table = data.sheets()[0]
    nrows = table.nrows
    nonstandard_addr_count = 0
    success_count = 0
    print("Finish reading xlsx")
    searched_list.seek(0,0)
    searched_list1=[line.encode('utf8').decode('utf-8-sig').replace("\n","") for line in searched_list]
    for i in range(nrows):
        num = table.row_values(i)[7] #路号
        street = table.row_values(i)[6] #路名
        if str(num).isspace() or street.isspace(): #如果路名或路号为空，则将名字和行政区域输入百度地图，返回相关条目作为此地址的补充
            nonstandard_addr_count = nonstandard_addr_count + 1
            keyword = table.row_values(i)[0] #公司/小区名字
            lat = table.row_values(i)[4]
            lng = table.row_values(i)[5]
            addr = table.row_values(i)[3]
            cur = str(keyword) + "^" + str(addr) + "^" + str(lat) + "^" + str(lng)
            # 如果是已爬取的信息，跳过
            if cur in searched_list1:
                print(cur+'已爬取')
                continue
            district = table.row_values(i)[2] #区名
            try:
                ak = BaiduAPI_singleSearch(keyword, district, ak)
                success_count = success_count + 1
                print("Finish  " + str(success_count) + " " + str(keyword))
                searched_list.write(cur + "\n")  # 写入txt
                searched_list.flush()
            except NoneAKException:
                sys.exit(0)
            except Exception as e:
                print(e)
                error_list.write(cur + "\n")  # 写入txt
                error_list.flush()
                continue
        else:
            pass
    print("============================================================================")
    print("Number of nonstandard address: " + str(nonstandard_addr_count))
    print("Number of successful search in BaiduAPI: " + str(success_count))
    print("============================================================================")


def txt_to_xlsx(filename, outfile): # txt转xlsx
    fr = open(filename, 'r+', encoding='utf8')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    row = 0
    for line in fr:
        row += 1
        line = line.strip()
        line = line.split('^')
        col = 0
        for j in range(len(line)):
            col += 1
            ws.cell(column=col, row=row, value=line[j].format(get_column_letter(col)))
    wb.save(outfile)


if __name__ == '__main__':
    global new_txt
    initial_AK_pond()
    global city_name_list
    global searched_list
    global searched_city
    global error_list
    city_name_list=['东营', '济南']
    searched_city=open("TEST/城市已爬列表.txt",'a+',encoding='utf8')
    searched_city.seek(0,0)
    searched_city_list=[line.encode('utf8').decode('utf-8-sig').replace("\n","") for line in searched_city]
    for city_name in city_name_list:
        if city_name in searched_city_list:
            print(city_name+'城市已经爬取')
            continue
        try:
            os.mkdir("TEST/%s"%(city_name))
        except OSError:
            pass

        new_txt = open("TEST/%s/%s公司地址补充.txt" %(city_name,city_name), 'a+', encoding='utf8')
        searched_list = open("TEST/%s/%s已爬列表.txt"%(city_name,city_name), 'a+', encoding='utf8')
        error_list = open("TEST/%s/%s报错列表.txt"%(city_name,city_name), 'a+', encoding='utf8')

        filepath = "/Users/gym/Desktop/Work/地址清洗项目/公司（正则后）/%s.xlsx" %(city_name)
        txt = 'TEST/%s/%s公司地址补充.txt' %(city_name,city_name)
        new_sheet = 'TEST/%s/%s公司地址补充.xlsx' %(city_name,city_name)
        run(filepath)
        print("Successfully saved to " + txt)
        new_txt.close()
        txt_to_xlsx(txt, new_sheet)
        print("Successfully converted to " + new_sheet)
        print("-------------------------------------ALL JOBS DONE-----------------------------------")
        searched_city.write(city_name+'\n')
        searched_city.flush()
        new_txt.close()
        searched_list.close()
        error_list.close()
    searched_city.close()