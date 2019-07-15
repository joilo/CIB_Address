# -*- coding:utf-8 -*-
import json
import openpyxl
from openpyxl.utils import get_column_letter
#import urllib
from urllib.request import urlopen, quote
import xlrd
import time
import os

class NoneAKException(Exception):
    def __init__(self,message):
        Exception.__init__(self)
        self.message=message

def initial_AK_pond():  # 初始化ak 池 0 为有额度 1为额度已经用完 保存成数组格式
    global ak_dic
    ak_dic = {}
    ak_dic = {
        "yHHqlqI0pLycBZVlMRjCFtQ8HSyMVWgv": 0,
        "xWvxINdrIy3G7awGlkBPulhDnY0NXA42": 0,
        "cON4dGCQZyz5IqglB7dYcbCPyrYxoxu1": 0,
    }

def exchange_AK():
    for line in ak_dic.items():
        if line[1] == 0:
            return line[0]
    print("ak池的额度全部用完了")
    return None

def BaiduAPI_singleSearch(master, key, region, ak):
    key_encode = quote(key)
    district = quote(region)
    url = "http://api.map.baidu.com/place/v2/search"
    url_send = url + "?query={}&region={}&output=json&ak={}".format(key_encode, district, ak)
    req = urlopen(url_send, timeout=60)
    res = req.read().decode()
    temp = json.loads(res)
    if temp['status'] == 0:
        name = key
        result = temp['results']
        for line in result:
            subname = line.get('name', '')
            city = line.get('city', '')
            area = line.get('area', '')
            addr = line.get('address', '')
            if 'location' in line:
                lat = line['location'].get('lat', '')
                lng = line['location'].get('lng', '')
            else:
                lat = ''
                lng = ''
            data = str(master) + "^" + str(name) + "^" + str(subname) + "^" + str(city) + "^" + str(area) + "^"\
                   + str(addr) + "^" + str(lat) + "^" + str(lng)
            new_txt.write(data + "\n")  # 写入txt
            #new_txt.flush()
    elif temp['status'] == 301 or temp['status'] == 302 or temp['status'] == 401 or temp['status'] == 402:
        ak_dic[ak] = 1  # 将当前AK的状态设置为已经跑完  P.S 1为已经跑完 0 为还有剩余额度
        print("捕获到AK额度不够的异常")
        ak = exchange_AK()  # 换一个AK
        print("已经更换AK", ak)
        if ak == None:  # 如果调用ak 之后为None 证明ak池的额度全部用完 错误文件记录当前运行结束时的状态
            print("配额全部用完啦！")
            raise NoneAKException("AK用完了")
        print("-----------------等待3s-------------------")
        time.sleep(3)
        BaiduAPI_singleSearch(master, key, region, ak)
    else:
        pass
    return ak


def run(filepath):
    ak = exchange_AK()
    data = xlrd.open_workbook(filepath)
    table = data.sheets()[0]
    nrows = table.nrows
    nonstandard_addr_count = 0
    success_count = 0
    print("Finish reading xlsx")
    searched_list.seek(0,0)
    searched_list1=[line.encode('utf8').decode('utf-8-sig').replace("\n","") for line in searched_list]
    for i in range(nrows):
        num = str(table.row_values(i)[8]) #路号
        street = str(table.row_values(i)[7]) #路名
        a = num.isspace()
        b = street.isspace()
        if a or b: #如果路名或路号为空，则将名字和行政区域输入百度地图，返回相关条目作为此地址的补充
            nonstandard_addr_count = nonstandard_addr_count + 1
            master = table.row_values(i)[0] #小区名字
            keyword = table.row_values(i)[1] #POI名字
            lat = table.row_values(i)[5]
            lng = table.row_values(i)[6]
            addr = table.row_values(i)[4]
            cur = str(keyword) + "^" + str(addr) + "^" + str(lat) + "^" + str(lng)
            # 如果是已爬取的信息，跳过
            if cur in searched_list1:
                print(cur+'已爬取')
                continue
            district = table.row_values(i)[3] #区名
            try:
                ak = BaiduAPI_singleSearch(master, keyword, district, ak)
                success_count = success_count + 1
                print("Finish  " + city + " " + str(success_count) + " " + str(keyword))
                searched_list.write(cur + "\n")  # 写入txt
                searched_list.flush()
            except Exception as e:
                print(e)
                error_list.write(cur + "\n")  # 写入txt
                error_list.flush()
                continue
        else:
            pass
    print("============================================================================")
    print("Number of nonstandard address: " + str(nonstandard_addr_count))
    print("Number of successful search in BaiduAPI: " + str(success_count))
    print("============================================================================")


def txt_to_xlsx(filename, outfile): # txt转xlsx
    fr = open(filename, 'r+', encoding='utf8')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    row = 0
    for line in fr:
        row += 1
        line = line.strip()
        line = line.split('^')
        col = 0
        for j in range(len(line)):
            col += 1
            ws.cell(column=col, row=row, value=line[j].format(get_column_letter(col)))
    wb.save(outfile)


if __name__ == '__main__':
    global new_txt
    initial_AK_pond()
    global city_name_list
    global searched_list
    global searched_city
    global error_list
    global city
    #city_name_list=['南平','南通','宁波','宁德','盘锦','平顶山','莆田','齐齐哈尔','钦州','青岛','曲靖']
    #ty_name_list=['泉州','日照','三明','三亚','厦门','汕头','上海','绍兴','深圳','沈阳',]
    #city_name_list=['十堰','石家庄','苏州','台州','太原','泰安',]
    city_name_list=['泰州','唐山','天津','威海','潍坊','渭南','温州','乌鲁木齐','无锡','芜湖','武汉','西安','西宁','咸阳','湘潭','襄阳',
                 '新乡','新余','信阳','邢台','宿迁','徐州','许昌','烟台','延边','盐城','扬州','伊犁','宜宾','宜昌','宜春','义乌','银川',
                 '鹰潭','营口','榆林','岳阳','湛江','漳州','长春','长沙','长治','镇江','郑州','中山','重庆','珠海','株洲','驻马店','遵义','淄博']
    searched_city=open("城市已爬列表.txt",'a+',encoding='utf8')
    searched_city.seek(0,0)
    searched_city_list=[line.encode('utf8').decode('utf-8-sig').replace("\n","") for line in searched_city]
    for city_name in city_name_list:
        city = city_name
        if city_name in searched_city_list:
            print(city_name+'城市已经爬取')
            continue
        os.mkdir("%s"%(city_name))
        new_txt = open("%s/%s小区地址补充.txt" %(city_name,city_name),'a+',encoding='utf8')
        searched_list = open("%s/%s已爬列表.txt"%(city_name,city_name),'a+',encoding='utf8')
        error_list = open("%s/%s报错列表.txt"%(city_name,city_name),'a+',encoding='utf8')
        filepath = "/Users/Shar/Desktop/地址项目/小区_二次迭代/%s_小区.xlsx" %(city_name)
        txt = '%s/%s小区地址补充.txt' %(city_name,city_name)
        new_sheet = '%s/%s小区地址补充.xlsx' %(city_name,city_name)
        run(filepath)
        print("Successfully saved to " + txt)
        new_txt.close()
        txt_to_xlsx(txt, new_sheet)
        print("Successfully converted to " + new_sheet)
        print("-------------------------------------ALL JOBS DONE-----------------------------------")
        searched_city.write(city_name+'\n')
        searched_city.flush()
        new_txt.close()
        searched_list.close()
        error_list.close()
    searched_city.close()
        