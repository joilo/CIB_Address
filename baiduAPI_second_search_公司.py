# -*- coding:utf-8 -*-
import json
import openpyxl
from openpyxl.utils import get_column_letter
# import urllib
from urllib.request import urlopen, quote
import xlrd
import time
import os


class NoneAKException(Exception):
    def __init__(self, message):
        Exception.__init__(self)
        self.message = message


def initial_AK_pond():  # 初始化ak 池 0 为有额度 1为额度已经用完 保存成数组格式
    global ak_dic
    ak_dic = {}
    ak_dic = {
        "mKOV9991B7N98C46kUeNRUXScM067Xys": 0,
        "g3LUhuGRHmgbf2SFwaCF8yAtGW4nezPS": 0,
        "kIzerPbQFAhs01H85XRNvngmXU73RF8L": 0,
        "hvZ5O0of19Alfl7HShUWyOVlFiD8WYWG": 0,
        "cq1NkGxVDIyVZSK1xLV2vsfVwOu5ajhV": 0,
        "5tlPS97zCcgTjbk6gy6AnLVG4p2jtg7u": 0,
        "UKg9gDjYcHMB5hSFMi1HxuQz18f041t1": 0,
        "GEC7Zek74HysO1AKCx1iG6bOXCzTWE6z": 0,
        "9HnBVwKEC01DMgxmINhOSGMt5q1M8kyr": 0,
        "SB3KV3mGWLQ3ncHEk7QfiRNCYHFMYtav": 0,
        "CDjmH9V1ZFfhv9qX2KzGvCrf8UVdUu99": 0,
        "y0j1hsIBRInpjyXub9dFLwsHjnTx73m4": 0,
        "YrwNoo8bNA2Nzfj7pldFXaXVz7iyEPXZ": 0,
        "jGDQ054Yx7n9MNFewodPgykG9UlvlYNa": 0,
        "c5AA39AFSpEAtDCRWWCRhoG5htUrUWvD": 0,
        "lN8FZ9Y8dXeGVdv7aTKvdcnkpXUuMcpQ": 0,
        "zer4hmUsf2Cppl2Z3ozkRMrGx6phGMVf": 0,
        "KdaCBLpAZrUApkiVqjYFSheusOwf2bhh": 0,
        "gwThbIBPPOlUYBQIMUhIP5haNLLkG3Nx": 0,
        "bKluUmlrV0daE5HMVMwgNbE7koRi8ur4": 0,
        "yHHqlqI0pLycBZVlMRjCFtQ8HSyMVWgv": 0,
        "xWvxINdrIy3G7awGlkBPulhDnY0NXA42": 0,
        "cON4dGCQZyz5IqglB7dYcbCPyrYxoxu1": 0,
    }


def exchange_AK():
    for line in ak_dic.items():
        if line[1] == 0:
            return line[0]
        else:
            continue
    print("ak池的额度全部用完了")
    return None

def mkdir(path):    #用于创建相对路径下的文件夹
    # os.makedirs(path)
    isExists=os.path.exists(path)

    if not isExists:
        # 如果不存在则创建目录
    # 创建目录操作函数
        os.makedirs(path)

        print(path + ' 创建成功')
        return True
    else:
        # 如果目录存在则不创建，并提示目录已存在
        print(path + ' 目录已存在')
        return False

def BaiduAPI_singleSearch(master, key, region, ak):
    key_encode = quote(key)
    district = quote(region)
    url = "http://api.map.baidu.com/place/v2/search"
    url_send = url + "?query={}&region={}&output=json&ak={}".format(key_encode, district, ak)
    req = urlopen(url_send, timeout=60)
    res = req.read().decode()
    temp = json.loads(res)
    if temp['status'] == 0:
        name = key
        result = temp['results']
        for line in result:
            subname = line.get('name', '')
            city = line.get('city', '')
            area = line.get('area', '')
            addr = line.get('address', '')
            if 'location' in line:
                lat = line['location'].get('lat', '')
                lng = line['location'].get('lng', '')
            else:
                lat = ''
                lng = ''
            data = str(master) + "^" + str(name) + "^" + str(subname) + "^" + str(city) + "^" + str(area) + "^" \
                   + str(addr) + "^" + str(lat) + "^" + str(lng)
            new_txt.write(data + "\n")  # 写入txt
            # new_txt.flush()
    elif temp['status'] == 301 or temp['status'] == 302 or temp['status'] == 401 or temp['status'] == 402:
        ak_dic[ak] = 1  # 将当前AK的状态设置为已经跑完  P.S 1为已经跑完 0 为还有剩余额度
        print("捕获到AK额度不够的异常")
        ak = exchange_AK()  # 换一个AK
        print("已经更换AK", ak)
        if ak == None:  # 如果调用ak 之后为None 证明ak池的额度全部用完 错误文件记录当前运行结束时的状态
            print("配额全部用完啦！")
            raise NoneAKException("AK用完了")
        print("-----------------等待3s-------------------")
        time.sleep(3)
        BaiduAPI_singleSearch(master=master, key=key, region=region, ak=ak)
    else:
        pass
    return ak

def cut_standard_add(xlsx_file, output_file):   #用于处理一个xlsx文件中的小区信息 并对其进行切分
    '''
    :param xlsx_file:  输入的Excel文件小区信息（未切分）
    :param output_file:  输出的txt文件（已切分）
    :return:
    '''
    fileObject = open(output_file, 'w', encoding='utf8')
    data = xlrd.open_workbook(xlsx_file)
    table = data.sheets()[0]
    nrows = table.nrows
    road_pattern = re.compile(".*?[^路]\((.*?)\)")
    road_pattern_1 = re.compile("(.*?(路|大道|街))")
    road_pattern_2 = re.compile("(.+)")
    num_leftpattern = re.compile(".*?(路|大道|街)(.+)")
    num_pattern = re.compile(".*?(\d+)")
    first_ten_pattern = re.compile("(十)[^一二三四五六七八九]")
    first_ten_1_pattern = re.compile("(十)[一二三四五六七八九]")
    middle_ten_pattern = re.compile(".*?[一二三四五六七八九](十)[一二三四五六七八九]")
    last_ten_pattern = re.compile(".*?[一二三四五六七八九](十)[^一二三四五六七八九]")
    for i in range(nrows):
        temp = str(table.row_values(i)[4]).strip()
        road_name = ' '
        road_match = road_pattern.match(temp)
        if road_match:
            road_name = road_match.group(1)
        else:
            road_match_1 = road_pattern_1.match(temp)
            if road_match_1:
                road_name = road_match_1.group(1)
            else:
                road_match_2 = road_pattern_2.match(temp)
                if road_match_2:
                    road_name = road_match_2.group(1)
        temp = str(table.row_values(i)[5]).strip()
        num_left = ' '
        num_leftmatch = num_leftpattern.match(temp)
        if num_leftmatch:
            num_left = num_leftmatch.group(2)
        num_str = ' '
        if num_left != ' ':
            first_ten = first_ten_pattern.match(num_left)
            first_ten_1 = first_ten_1_pattern.match(num_left)
            middle_ten = middle_ten_pattern.match(num_left)
            last_ten = last_ten_pattern.match(num_left)

            if first_ten:
                num_left = num_left.replace("十", "10")
            elif first_ten_1:
                num_left = num_left.replace("十", "1")
            elif middle_ten:
                num_left = num_left.replace("十", "")
            elif last_ten:
                num_left = num_left.replace("十", "0")
            num_left = num_left.replace("一", "1")
            num_left = num_left.replace("二", "2")
            num_left = num_left.replace("三", "3")
            num_left = num_left.replace("四", "4")
            num_left = num_left.replace("五", "5")
            num_left = num_left.replace("六", "6")
            num_left = num_left.replace("七", "7")
            num_left = num_left.replace("八", "8")
            num_left = num_left.replace("九", "9")
            match_num = num_pattern.match(num_left)
            if match_num:
                num_str = match_num.group(1)
        fileObject.write(str(table.row_values(i)[0]).strip() + '^' + str(table.row_values(i)[1]).strip() + '^' +
                        str(table.row_values(i)[2]).strip() + '^' + str(table.row_values(i)[3]).strip() + '^' +
                        str(table.row_values(i)[5]).strip() + '^' + str(table.row_values(i)[6]).strip() + '^' +
                        str(table.row_values(i)[7]).strip() + '^' + road_name + '^' + num_str + '\n')
    fileObject.close()

def reverseLng(lng, lat,ak):   #经纬度反向解析, 经度在前 纬度在后  获得标准地址库所用json文件
    add_list = []
    url = "http://api.map.baidu.com/geocoder/v2/"
    output = 'json'
    ak = ak
    url_send = url + "?callback=renderReverse&location=%s,%s&output=json&pois=1&latest_admin=1&ak=%s&radius=300" % (lat, lng, ak)
    req = urlopen(url_send)
    res = req.read().decode()  # 将其他编码的字符串解码成unicode
    if str(res[:29]) == 'renderReverse&&renderReverse(':
        temp = json.loads(res[29:-1])
        city = temp['result']['addressComponent']['city']
        district = temp['result']['addressComponent']['district']
        formatted_address = temp['result']['formatted_address']
        pois = temp['result']['pois']
        street = temp['result']['addressComponent']['street']  # 获取小区所在街道 用于建立标准路库
        for line in pois:
            model = []
            model.append(line['addr'])
            model.append(line['distance'])
            model.append(line['name'])
            lng = str(line['point']['x'])
            lat = str(line['point']['y'])
            model.append(lat)
            model.append(lng)
            add_list.append(model)
        return add_list, city, district, formatted_address, street
    else:
        temp = json.loads(res)
        if temp['status'] == 301 or temp['status'] == 302 or temp['status'] == 401 or temp['status'] == 402:
            ak_dic[ak] = 1  # 将当前AK的状态设置为已经跑完  P.S 1为已经跑完 0 为还有剩余额度
            print("捕获到AK额度不够的异常")
            for i in range(len(ak_dic)):
                ak_dic[ak] = 1
                ak = exchange_AK()  # 换一个AK
                print("已经更换AK", ak)
                if ak == None:  # 如果调用ak 之后为None 证明ak池的额度全部用完 错误文件记录当前运行结束时的状态
                    print("配额全部用完啦！")
                    raise NoneAKException("AK用完了")
                print("-----------------等待3s-------------------")
                time.sleep(3)
                reverseLng(lng, lat, ak)

def run(filepath):
    ak = exchange_AK()
    data = xlrd.open_workbook(filepath)
    table = data.sheets()[0]
    nrows = table.nrows
    nonstandard_addr_count = 0
    success_count = 0
    print("Finish reading xlsx")
    searched_list.seek(0, 0)
    searched_list1 = [line.encode('utf8').decode('utf-8-sig').replace("\n", "") for line in searched_list]
    for i in range(nrows):
        num = str(table.row_values(i)[7])  # 路号
        street = str(table.row_values(i)[6])  # 路名 注意这里和小区的格式不一样
        if str(num).isspace() or street.isspace():  # 如果路名或路号为空，则将名字和行政区域输入百度地图，返回相关条目作为此地址的补充
            nonstandard_addr_count = nonstandard_addr_count + 1
            master = table.row_values(i)[0]  # 小区名字
            city = table.row_values(i)[1]
            district = table.row_values(i)[2]
            lat = table.row_values(i)[4]
            lng = table.row_values(i)[5]
            addr = table.row_values(i)[3]
            cur = str(master) + "^" + str(addr) + "^" + str(lat) + "^" + str(lng)
            # 如果是已爬取的信息，跳过
            if cur in searched_list1:
                print(cur + '已爬取')
                continue
            district = table.row_values(i)[3]  # 区名
            try:
                add_list, city, district, formatted_address, street = reverseLng(lng, lat, ak)
                data = str(master) + "^" + str(master) + "^" + str(city) + "^" + str(district) + "^" + str(street) + "^" \
                       + str(formatted_address) + "^" + str(lat) + "^" + str(lng)
                new_txt.write(data + "\n")
                for line in add_list:  #将POI 逆向解析
                    name = line[2]
                    formatted_address = line[0]
                    lat = line[3]
                    lng = line[4]
                    add_list, city, district, formatted_address, street = reverseLng(lng, lat, ak)
                    data = str(master) + "^" + str(name) + "^" + str(city) + "^" + str(district) + "^" + str(
                        street) + "^" \
                           + str(formatted_address) + "^" + str(lat) + "^" + str(lng)
                    new_txt.write(data + "\n")
                new_txt.flush()
                success_count = success_count + 1
                print("Finish  " + str(success_count) + " " + str(master))
                searched_list.write(cur + "\n")  # 写入txt
                searched_list.flush()
            except Exception as e:
                print(e)
                error_list.write(cur + "\n")  # 写入txt
                error_list.flush()
                continue
        else:
            pass
    print("============================================================================")
    print("Number of nonstandard address: " + str(nonstandard_addr_count))
    print("Number of successful search in BaiduAPI: " + str(success_count))
    print("============================================================================")


def txt_to_xlsx(filename, outfile):  # txt转xlsx
    fr = open(filename, 'r+', encoding='utf8')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    row = 0
    for line in fr:
        row += 1
        line = line.strip()
        line = line.split('^')
        col = 0
        for j in range(len(line)):
            col += 1
            ws.cell(column=col, row=row, value=line[j].format(get_column_letter(col)))
    wb.save(outfile)


if __name__ == '__main__':
    global new_txt
    initial_AK_pond()
    global city_name_list
    global searched_list
    global searched_city
    global error_list
    path = './公司地址第三次迭代'
    mkdir(path)
    city_name_list = ['黄石']
    searched_city = open("./公司地址第三次迭代/城市已爬列表.txt", 'a+', encoding='utf8')
    searched_city.seek(0, 0)
    searched_city_list = [line.encode('utf8').decode('utf-8-sig').replace("\n", "") for line in searched_city]
    print(searched_city_list)
    for city_name in city_name_list:
        if city_name in searched_city_list:
            print(city_name + '城市已经爬取')
            continue
        path = "./公司地址第三次迭代/%s" % (city_name)
        mkdir(path)
        new_txt = open("./公司地址第三次迭代/%s/%s_公司地址补充.txt" % (city_name, city_name), 'a+', encoding='utf8')
        searched_list = open("./公司地址第三次迭代/%s/%s已爬列表.txt" % (city_name, city_name), 'a+', encoding='utf8')
        error_list = open("./公司地址第三次迭代/%s/%s报错列表.txt" % (city_name, city_name), 'a+', encoding='utf8')
        filepath = "D:\\working\\python\\XingYe\\CIB_Address\\公司+小区标准路（正则后）\\公司（正则后）\\%s.xlsx" % (city_name)
        txt = './公司地址第三次迭代/%s/%s_公司地址补充.txt' % (city_name, city_name)
        new_sheet = './公司地址第三次迭代/%s/%s_公司地址补充.xlsx' % (city_name, city_name)
        run(filepath)
        print("Successfully saved to " + txt)
        new_txt.close()
        txt_to_xlsx(txt, new_sheet)
        print("Successfully converted to " + new_sheet)
        outfile_road = "./公司地址第三次迭代/%s/%s公司地址补充_切分.txt" % (city_name,city_name)
        outfile_road1 = "./公司地址第三次迭代/%s/%s公司地址补充_切分.xlsx" % (city_name, city_name)
        cut_standard_add(new_sheet,outfile_road)
        print(city_name, "切分已完成")
        txt_to_xlsx(outfile_road,outfile_road1)
        print("============================转化xlsx文件完成==========================================")
        print("-------------------------------------ALL JOBS DONE-----------------------------------")
        searched_city.write(city_name + '\n')
        searched_city.flush()
        new_txt.close()
        searched_list.close()
        error_list.close()
    searched_city.close()
